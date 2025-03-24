import os
import re
import json
import hashlib
import logging
from datetime import datetime
from flask import Flask, request, jsonify

from openpyxl import load_workbook
from google.cloud import pubsub_v1

logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s | %(asctime)s | %(name)s | %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

def converter_data_br(valor):
    """Converte datas em formato brasileiro (dd/mm/aaaa) para iso (yyyy-mm-dd)."""
    if isinstance(valor, str):
        valor = valor.strip()
        for fmt in ["%d/%m/%Y", "%d/%m/%y"]:
            try:
                dt = datetime.strptime(valor, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                pass
    return valor

def converter_valor_br(valor):

    if isinstance(valor, str):
        limpo = re.sub(r"[R$\s]", "", valor)
        limpo = limpo.replace(".", "")  
        limpo = limpo.replace(",", ".") 
        try:
            return float(limpo)
        except ValueError:
            return None
    return valor

def compute_row_hash(row, current_columns, account):
    """
    Gera um hash MD5 de cada linha baseado em colunas relevantes + account.
    """
    parts = []
    parts.append(str(account))
    for col in current_columns:
        if col is not None:
            parts.append(str(row.get(col, "")))
    concatenated = "".join(parts)
    return hashlib.md5(concatenated.encode('utf-8')).hexdigest()

def convert_data(file_path, account):

    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    data_blocks = []
    current_block = []
    current_columns = []
    
    STATE_SEARCHING_HEADER = 0
    STATE_READING_DATA     = 1
    STATE_SEARCHING_NEXT   = 2

    state = STATE_SEARCHING_HEADER
    lines_without_header = 0
    
    for row in sheet.iter_rows(values_only=True):
        first_cell = row[0] if row else None
        if isinstance(first_cell, str):
            first_cell = first_cell.strip().lower()
        if not first_cell:
            first_cell = None
        
        if state == STATE_SEARCHING_HEADER:
            if first_cell == 'data':
                current_columns = [
                    str(x).strip().lower() if x else None 
                    for x in row
                ]
                current_block = []
                state = STATE_READING_DATA
            else:
                continue
        
        elif state == STATE_READING_DATA:
            if first_cell is None:
                # fim do bloco
                if current_block:
                    data_blocks.append(current_block)
                current_block = []
                current_columns = []
                state = STATE_SEARCHING_NEXT
                lines_without_header = 0
            else:
                row_dict = {}
                for col_index, col_name in enumerate(current_columns):
                    if col_name:
                        valor_celula = row[col_index] if col_index < len(row) else None
                        row_dict[col_name] = valor_celula
                # conversões
                if 'data' in row_dict:
                    row_dict['data'] = converter_data_br(row_dict['data'])
                if 'valor' in row_dict:
                    row_dict['valor'] = converter_valor_br(row_dict['valor'])
                
                row_dict['account'] = account
                row_dict['id'] = compute_row_hash(row_dict, current_columns, account)
                
                current_block.append(row_dict)
        
        elif state == STATE_SEARCHING_NEXT:
            if first_cell == 'data':
                current_columns = [
                    str(x).strip().lower() if x else None 
                    for x in row
                ]
                current_block = []
                state = STATE_READING_DATA
            else:
                lines_without_header += 1
                if lines_without_header >= 6:
                    break
                else:
                    continue

    # se terminar o arquivo ainda no estado de leitura de dados
    if state == STATE_READING_DATA and current_block:
        data_blocks.append(current_block)
    
    return data_blocks

@app.route('/', methods=['POST'])
def parse_excel():
    """
    Função HTTP que:
      - Recebe via JSON um 'file_path' e um 'account' (ou assume default).
      - Lê o Excel e separa em blocos (listas).
      - Publica cada linha de cada bloco no Pub/Sub.
    Retorna um JSON com contagem de mensagens publicadas, e logs no Cloud Logging.
    """
    try:
        request_json = request.get_json(silent=True)
        if not request_json:
            logger.warning("Nenhum JSON na requisição.")
            return jsonify({"error": "JSON inválido ou inexistente"}), 400

        file_path = request_json.get("file_path")
        account   = request_json.get("account", "Cartão Azul Visa")

        if not file_path:
            logger.warning("Parâmetro 'file_path' não informado.")
            return jsonify({"error": "Parâmetro 'file_path' é obrigatório"}), 400
        
        logger.info(f"Iniciando parse do arquivo: {file_path} para account={account}.")
        
        data_blocks = convert_data(file_path, account)
        total_blocos = len(data_blocks)
        logger.info(f"Foram identificados {total_blocos} bloco(s) no arquivo.")

        # Configura Pub/Sub
        publisher = pubsub_v1.PublisherClient()
        topic_path = os.environ.get("TRANSACTIONS_TOPIC")

        if not topic_path:
            msg = "A variável de ambiente 'TRANSACTIONS_TOPIC' não foi definida."
            logger.error(msg)
            return jsonify({"error": msg}), 500
        
        total_mensagens = 0

        # Publica cada linha de cada bloco no Pub/Sub
        for i, bloco in enumerate(data_blocks, start=1):
            logger.info(f"Publicando bloco {i} com {len(bloco)} linhas.")
            for row_dict in bloco:
                try:
                    row_json = json.dumps(row_dict).encode("utf-8")
                    future = publisher.publish(topic_path, data=row_json)
                    message_id = future.result()
                    total_mensagens += 1
                except Exception as e:
                    logger.error(f"Erro ao publicar no Pub/Sub: {e}", exc_info=True)
                    # dependendo da lógica, você poderia continuar ou retornar erro
                    # vamos continuar
                    continue
        
        logger.info(f"Publicação concluída. Total de mensagens publicadas: {total_mensagens}.")

        return jsonify({
            "file_path": file_path,
            "account": account,
            "blocos": total_blocos,
            "mensagens_publicadas": total_mensagens
        }), 200
    
    except Exception as e:
        logger.error(f"Erro inesperado na função: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

def main(request):
    """Ponto de entrada para Cloud Functions (HTTP)."""
    return app(request)
