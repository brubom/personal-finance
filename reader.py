import re
import os
import json
from datetime import datetime
from openpyxl import load_workbook
import hashlib
from google.cloud import pubsub_v1
from flask import Flask

app = Flask(__name__)

def converter_data_br(valor):
    if isinstance(valor, str):
        valor = valor.strip()
        for fmt in ["%d/%m/%Y", "%d/%m/%y"]:
            try:
                dt = datetime.strptime(valor, fmt)
                return dt.strftime("%Y-%m-%d")  
            except ValueError:
                pass
    return valor

def converter_valor_br(valor):
    if isinstance(valor, str):
        limpo = re.sub(r"[R$\s]", "", valor)
        limpo = limpo.replace(".", "")  
        limpo = limpo.replace(",", ".") 
        try:
            return float(limpo)
        except ValueError:
            return None
    return valor

def compute_row_hash(row, current_columns, account):

    parts = []
    parts.append(str(account))
    for col in current_columns:
        if col is not None:
            parts.append(str(row.get(col, "")))
    concatenated = "".join(parts)
    return hashlib.md5(concatenated.encode('utf-8')).hexdigest()

def convert_data(file_path, account):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    data_blocks = []       
    current_block = []     
    current_columns = []   
    
    # Definindo estados
    STATE_SEARCHING_HEADER = 0
    STATE_READING_DATA     = 1
    STATE_SEARCHING_NEXT   = 2

    state = STATE_SEARCHING_HEADER
    lines_without_header = 0  
    
    for row in sheet.iter_rows(values_only=True):
        first_cell = row[0] if row else None
        
        if isinstance(first_cell, str):
            first_cell = first_cell.strip().lower()
        if not first_cell:
            first_cell = None
        
        if state == STATE_SEARCHING_HEADER:
            if first_cell == 'data':
                current_columns = [str(x).strip().lower() if x else None for x in row]
                current_block = []
                state = STATE_READING_DATA
            else:
                continue  
        
        elif state == STATE_READING_DATA:
            if first_cell is None:
                if current_block:
                    data_blocks.append(current_block)
                current_block = []
                current_columns = []
                state = STATE_SEARCHING_NEXT
                lines_without_header = 0
            else:
                row_dict = {}
                for col_index, col_name in enumerate(current_columns):
                    if col_name:
                        valor_celula = row[col_index] if col_index < len(row) else None
                        row_dict[col_name] = valor_celula
                if 'data' in row_dict:
                    row_dict['data'] = converter_data_br(row_dict['data'])
                if 'valor' in row_dict:
                    row_dict['valor'] = converter_valor_br(row_dict['valor'])
                row_dict['account'] = account
                row_dict['id'] = compute_row_hash(row_dict, current_columns, account)
                
                current_block.append(row_dict)
        
        elif state == STATE_SEARCHING_NEXT:
            if first_cell == 'data':
                current_columns = [str(x).strip().lower() if x else None for x in row]
                current_block = []
                state = STATE_READING_DATA
            else:
                lines_without_header += 1
                if lines_without_header >= 6:
                    break  
                else:
                    continue

    if state == STATE_READING_DATA and current_block:
        data_blocks.append(current_block)
    
    return data_blocks

@app.route('/', methods=['POST'])
def parse_excel(file_path):
    data = convert_data(file_path, "Cartão Azul Visa")

    publisher = pubsub_v1.PublisherClient()
    topic_path = os.environ.get("transactions_topic")

    for row in data:
        row_json = json.dumps(row).encode("utf-8")
        future = publisher.publish(topic_path, data=row_json)
        message_id = future.result()
        
    

if __name__ == "__main__":
    file_path = "Fatura-Excel-fev.xlsx"  
    account = "Cartão Azul Visa"         
    
    blocks = convert_data(file_path, account)
    
    print(f"Foram encontrados {len(blocks)} bloco(s) de dados.\n")
    for i, block in enumerate(blocks, start=1):
        print(f"--- Bloco {i} (exibindo até 5 linhas) ---")
        for row_data in block[:5]:
            print(row_data)