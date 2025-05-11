import os
import re
import json
import hashlib
import logging
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from typing import List, Dict, Any, Tuple
import time

import functions_framework
from flask import Request
from utils.telemetry import create_span, get_current_trace_id
from utils.factories import get_logger, get_telemetry, get_pubsub_publisher, get_topic_path

# Setup logger
logger = get_logger(__name__)

def converter_data_br(data_str: str) -> str:
    """Converte data do formato brasileiro para ISO."""
    try:
        if not isinstance(data_str, str):
            return data_str
        
        # Tenta converter data no formato dd/mm/yyyy
        try:
            data = datetime.strptime(data_str, "%d/%m/%Y")
        except ValueError:
            # Tenta converter data no formato dd/mm/yy
            data = datetime.strptime(data_str, "%d/%m/%y")
        
        return data.strftime("%Y-%m-%d")
    except Exception:
        return data_str

def converter_valor_br(valor_str: str) -> float:
    """Converte valor do formato brasileiro para float."""
    try:
        if not isinstance(valor_str, str):
            return valor_str
        
        # Remove R$ e espaços
        valor_str = valor_str.replace('R$', '').strip()
        
        # Converte para float
        return float(valor_str.replace('.', '').replace(',', '.'))
    except Exception:
        return None

def compute_row_hash(row: Dict[str, Any], columns: List[str], account: str) -> str:
    """Computa hash MD5 para uma linha de dados."""
    # Cria string com valores concatenados
    values = [str(row.get(col, '')) for col in columns]
    values.append(account)
    data = ''.join(values)
    
    # Computa hash MD5
    return hashlib.md5(data.encode()).hexdigest()

def convert_xls_to_xlsx(file_path: str) -> str:
    """Converte arquivo .xls para .xlsx."""
    with create_span("convert_xls_to_xlsx", {"file_path": file_path}) as span:
        try:
            logger.info("Converting XLS to XLSX", extra={"file_path": file_path})
            
            # Lê o arquivo .xls
            df = pd.read_excel(file_path, engine='xlrd')
            
            # Cria novo caminho para arquivo .xlsx
            xlsx_path = file_path.replace('.xls', '.xlsx')
            
            # Salva como .xlsx
            df.to_excel(xlsx_path, index=False)
            
            logger.info("XLS to XLSX conversion completed",
                          extra={"original_file": file_path,
                                 "converted_file": xlsx_path})
            
            return xlsx_path
        except Exception as e:
            error_msg = f"Error converting XLS to XLSX: {str(e)}"
            logger.error(error_msg, extra={"file_path": file_path})
            span.set_attribute("error", error_msg)
            raise

def process_row(row, current_columns, account):
    """Processa uma linha de dados."""
    row_dict = {}
    for col_index, col_name in enumerate(current_columns):
        if col_name:
            valor_celula = row[col_index] if col_index < len(row) else None
            row_dict[col_name] = valor_celula
    
    if 'data' in row_dict:
        row_dict['data'] = converter_data_br(row_dict['data'])
    if 'valor' in row_dict:
        row_dict['valor'] = converter_valor_br(row_dict['valor'])
    
    row_dict['account'] = account
    row_dict['id'] = compute_row_hash(row_dict, current_columns, account)
    return row_dict

def process_header(row):
    """Processa o cabeçalho da planilha."""
    return [str(x).strip().lower() if x else None for x in row]

def convert_data(file_path: str, account: str) -> List[List[Dict[str, Any]]]:
    """Converte dados do Excel para lista de dicionários."""
    with create_span("convert_data", {"file_path": file_path, "account": account}) as span:
        try:
            logger.info("Starting data conversion", extra={"file_path": file_path, "account": account})
            
            # Converte .xls para .xlsx se necessário
            if file_path.endswith('.xls'):
                file_path = convert_xls_to_xlsx(file_path)
            
            # Carrega workbook
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            # Lista para armazenar blocos de dados
            data_blocks = []
            current_block = []
            
            # Processa linhas
            for row in sheet.iter_rows(values_only=True):
                # Verifica se é uma linha de cabeçalho
                if row[0] == 'data':
                    # Se já temos um bloco, adiciona à lista
                    if current_block:
                        data_blocks.append(current_block)
                        current_block = []
                    continue
                
                # Verifica se é uma linha vazia
                if not any(row):
                    if current_block:
                        data_blocks.append(current_block)
                        current_block = []
                    continue
                
                # Processa linha de dados
                data = converter_data_br(row[0])
                valor = converter_valor_br(row[1])
                descricao = row[2]
                
                # Cria dicionário com dados
                row_data = {
                    'data': data,
                    'valor': valor,
                    'descricao': descricao,
                    'account': account
                }
                
                # Adiciona hash
                row_data['id'] = compute_row_hash(
                    row_data,
                    ['data', 'valor', 'descricao'],
                    account
                )
                
                current_block.append(row_data)
            
            # Adiciona último bloco se houver
            if current_block:
                data_blocks.append(current_block)
            
            logger.info("Data conversion completed",
                          extra={"file_path": file_path,
                                 "blocks_count": len(data_blocks),
                                 "total_rows": sum(len(block) for block in data_blocks)})
            
            return data_blocks
        except Exception as e:
            error_msg = f"Error processing file: {str(e)}"
            logger.error(error_msg, extra={"file_path": file_path})
            span.set_attribute("error", error_msg)
            raise

@functions_framework.http
def parse_excel(request: Request, publisher=None, topic_path=None, telemetry=None):
    """HTTP Cloud Function para processar arquivo Excel."""
    publisher = publisher or get_pubsub_publisher()
    topic_path = topic_path or get_topic_path(publisher)
    telemetry = telemetry or get_telemetry("azul_visa_reader")
    
    with create_span("parse_excel") as span:
        try:
            # Verificar se é uma requisição local
            if not os.getenv("K_SERVICE"):
                return "Running in local mode"
            
            # Obter dados da requisição
            request_json = request.get_json(silent=True)
            if not request_json:
                error_msg = "Invalid JSON in request"
                logger.error(error_msg)
                span.set_attribute("error", error_msg)
                return (error_msg, 400)
            
            file_path = request_json.get("file_path")
            if not file_path:
                error_msg = "Missing file_path in request"
                logger.error(error_msg)
                span.set_attribute("error", error_msg)
                return (error_msg, 400)
            
            # Registrar início do processamento
            logger.info("Starting file processing", extra={"file_path": file_path})
            
            # Converter XLS para XLSX se necessário
            if file_path.endswith(".xls"):
                start_time = time.monotonic()
                file_path = convert_xls_to_xlsx(file_path)
                conversion_duration = time.monotonic() - start_time
                logger.info("SLI: xls_conversion_duration",
                          extra={"duration": conversion_duration, "file_path": file_path})
            
            # Ler dados do Excel
            start_time = time.monotonic()
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Processar linhas
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Verificar se a linha não está vazia
                    rows.append({
                        "data": row[0],
                        "descricao": row[1],
                        "valor": row[2]
                    })
            
            # Converter dados
            converted_rows = convert_data(file_path, 'ITAU_CARD')
            
            # Publicar mensagem no Pub/Sub
            message = {
                "rows": converted_rows,
                "file_path": file_path,
                "trace_id": get_current_trace_id()
            }
            
            future = publisher.publish(
                topic_path,
                json.dumps(message).encode("utf-8")
            )
            future.result()
            
            # Registrar métricas
            processing_duration = time.monotonic() - start_time
            logger.info("SLI: processing_duration",
                       extra={"duration": processing_duration,
                             "rows_processed": len(converted_rows),
                             "file_path": file_path})
            
            # Registrar sucesso
            logger.info("File processed successfully",
                       extra={"file_path": file_path,
                             "rows_processed": len(converted_rows)})
            
            return ("OK", 200)
            
        except Exception as e:
            error_msg = f"Error processing file: {str(e)}"
            logger.error(error_msg,
                        extra={"file_path": file_path if 'file_path' in locals() else None})
            span.set_attribute("error", error_msg)
            return (error_msg, 500)

def main(logger=None, telemetry=None):
    logger = logger or get_logger(__name__)
    telemetry = telemetry or get_telemetry("writer")
    # ... resto do código ...

from unittest.mock import MagicMock

def test_main():
    fake_logger = MagicMock()
    fake_telemetry = MagicMock()
    main(logger=fake_logger, telemetry=fake_telemetry)
    # ... asserts ...
